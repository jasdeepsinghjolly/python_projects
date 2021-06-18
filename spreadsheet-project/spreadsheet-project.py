import openpyxl      # importing module openpyxl for spreadsheet functions
excel_file = openpyxl.load_workbook("inventory.xlsx")    # reading complete workbook and saving it in the variable
# excel_file
print(type(excel_file))     # O/P : <class 'openpyxl.workbook.workbook.Workbook'>
product_sheet = excel_file["Sheet1"]   # extracting only sheet "Sheet1" from excel and saving it in the variable.
print(type(product_sheet))  # O/P : <class 'openpyxl.worksheet.worksheet.Worksheet'>
print(product_sheet.max_row)  # O/P : 75
print(range(product_sheet.max_row))  # O/P : range(0, 75) ; range function provide the range for number given.
print(range(2, product_sheet.max_row + 1))  # O/P : range(2, 75) ; In this we have asked range() function to start its
# range from 2 until 75 because as per our data in the "Sheet1" we want to ignore the headline (i.e 1st row) in this sheet.
product_per_supplier = {}  # creating an empty dictionary
total_value_per_supplier = {}
products_under_10_inv = {}
for product_row in range(2, product_sheet.max_row + 1):  # "FOR" loop is meant to used for iterating over a list; now in
# this case range function will provide list of number from 0 to 74 for "FOR" loop to iterate over. By this
# we will be having a valid "FOR" loop of from 0 to 74 and then FOR loop will iterate over all rows present in the sheet.
# Also "+1" is added into last digit as we want to include row 75 also to be included into our range as range()
# function does not consider last number ; Now are range would be from 2 to 75 (with both start and end numbers included).
    supplier_name = product_sheet.cell(product_row, 4).value  # In this we are saving value from column "supplier" from the sheet
    # i.e column 4 in a "supplier_name" variable.For this we are using cell() function on our "Sheet1" sheet and
    # providing row number ( which is "FOR" loop is iterating )and column 4 ( for supplier name from the sheet)
    # to cell() as parameter.cell() is the function used to get the value from the cell by providing row and column number to it.
    #print(supplier_name)
    inventory = product_sheet.cell(product_row, 2).value
    price = product_sheet.cell(product_row, 3).value
    product_number = product_sheet.cell(product_row, 1).value
    inventory_price = product_sheet.cell(product_row, 5)

# Calculation to get total count for each supplier
    if supplier_name in product_per_supplier:      # To check fs the suppliers name is already added to the dictionary then
        # increament the value corrosponding the supplier name in the dictionary in the next 2 lines.
       current_num_product = product_per_supplier.get(supplier_name)  # In this Value of the key (supplier name i.e AAA company)
       # from dictionary product_per_supplier is added to variable (current_num_product)
       product_per_supplier[supplier_name] = current_num_product +1   # Increamenting the count on every iteration of "FOR" loop
       # and adding it to the previous value of the key suppliers name in dictionary "product_per_supplier". Both these
       # lines can be replaced with "product_per_supplier[supplier_name] = product_per_supplier[supplier_name] +1"
    else:
        # print(f"Adding a New Supplier {supplier_name} for the first time")
        product_per_supplier[supplier_name] = 1  # This will be executed only for the first time when we encounter a new supplier name
        # which is not already present in our "product_per_supplier" dictionary and it will put value "1" for the key
        # supplier_name (AAA comapany etc) in the dictionary.

# Calculation for getting value of total inventories for each supplier i.e inventory * price.
    if supplier_name in total_value_per_supplier:
        current_total_value = total_value_per_supplier.get(supplier_name)
        total_value_per_supplier[supplier_name] = current_total_value + (inventory * price)
    else:
        print(f"Adding a New Supplier {supplier_name} for the first time")
        total_value_per_supplier[supplier_name] = inventory * price

# Calculation to get product number that have inventory under 10
    if inventory < 10:
       products_under_10_inv[product_number] = inventory

# Add column into sheet for inventory price.
    inventory_price.value = inventory * price  # We are updating the value for column 5 for each row in the sheet with
    # multiplication of inventory value and price value. But this will save the values in temporary file but not
    # actually not save on sheet.
excel_file.save("inventory_with_total_value.xlsx")   # We are saving the complete workbook ,
    # in this we will not overwrite the existing file but create a new workbook with name "inventory_with_total_value.xlsx"
    # using save() function from the openpyxl module.
print(product_per_supplier)
print(total_value_per_supplier)
print(products_under_10_inv)